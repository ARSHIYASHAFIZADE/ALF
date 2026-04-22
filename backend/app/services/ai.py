"""AI-powered file analysis — Groq-hosted Llama.

Single entry point: ``analyze_job(job)`` pulls a short preview out of the
user's file (text for docs, metadata for media, listing for archives, etc.),
sends it to the model, and returns a structured insight object the UI can
render directly.
"""
from __future__ import annotations

import json
import logging
import subprocess
import zipfile
from pathlib import Path
from typing import Optional

from groq import Groq

from app.config import settings

log = logging.getLogger("ai")

_client: Optional[Groq] = None


def _get_client() -> Optional[Groq]:
    global _client
    if not settings.GROQ_API_KEY:
        return None
    if _client is None:
        _client = Groq(api_key=settings.GROQ_API_KEY)
    return _client


# ---------- content extractors ----------

MAX_PREVIEW_CHARS = 3500


def _read_text(path: Path, limit: int = MAX_PREVIEW_CHARS) -> str:
    try:
        data = path.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        return f"(unable to read as text: {e})"
    if len(data) > limit:
        data = data[:limit] + f"\n… (truncated, full size {path.stat().st_size} bytes)"
    return data


def _preview_document(path: Path, fmt: str) -> str:
    if fmt == "pdf":
        try:
            import fitz
            doc = fitz.open(str(path))
            out = [f"PDF · {len(doc)} page(s)"]
            for i, page in enumerate(doc):
                if i >= 3:
                    break
                txt = page.get_text("text").strip()
                if txt:
                    out.append(f"--- page {i + 1} ---\n{txt[:1500]}")
            doc.close()
            return "\n".join(out)[:MAX_PREVIEW_CHARS]
        except Exception as e:
            return f"(pdf preview failed: {e})"
    if fmt == "docx":
        try:
            from docx import Document
            d = Document(str(path))
            paras = [p.text for p in d.paragraphs if p.text.strip()]
            head = paras[:25]
            return f"DOCX · {len(paras)} paragraph(s)\n\n" + "\n".join(head)[:MAX_PREVIEW_CHARS]
        except Exception as e:
            return f"(docx preview failed: {e})"
    if fmt in ("xlsx", "xls"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)
            lines = []
            for sheet in wb.sheetnames[:3]:
                ws = wb[sheet]
                lines.append(f"--- sheet {sheet} ({ws.max_row}×{ws.max_column}) ---")
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 10:
                        break
                    lines.append(" | ".join(str(c) if c is not None else "" for c in row))
            return "\n".join(lines)[:MAX_PREVIEW_CHARS]
        except Exception as e:
            return f"(xlsx preview failed: {e})"
    # txt / md / html / csv / tsv / rtf — all textual
    return _read_text(path)


def _preview_image(path: Path) -> str:
    try:
        from PIL import Image, ExifTags
        img = Image.open(str(path))
        info = [
            f"Image · {img.format}",
            f"Size: {img.width}×{img.height} px",
            f"Mode: {img.mode}",
        ]
        # EXIF where available
        try:
            exif = img.getexif()
            named = {ExifTags.TAGS.get(k, k): v for k, v in exif.items() if k in ExifTags.TAGS}
            interesting = {
                k: named[k]
                for k in ("Make", "Model", "DateTime", "ExposureTime", "ISOSpeedRatings", "FNumber", "FocalLength", "GPSInfo")
                if k in named
            }
            if interesting:
                info.append("EXIF: " + ", ".join(f"{k}={v}" for k, v in interesting.items()))
        except Exception:
            pass
        return "\n".join(info)
    except Exception as e:
        return f"(image preview failed: {e})"


def _preview_media(path: Path) -> str:
    try:
        r = subprocess.run(
            ["ffprobe", "-v", "error", "-show_format", "-show_streams", "-print_format", "json", str(path)],
            capture_output=True, text=True, timeout=15,
        )
        if r.returncode != 0:
            return f"(ffprobe failed: {r.stderr[-200:]})"
        data = json.loads(r.stdout or "{}")
        fmt = data.get("format", {})
        streams = data.get("streams", [])
        lines = [
            f"Container: {fmt.get('format_name', '?')}",
            f"Duration: {float(fmt.get('duration', 0)):.1f}s",
            f"Bitrate: {int(fmt.get('bit_rate', 0)) // 1000} kbps" if fmt.get("bit_rate") else "Bitrate: ?",
        ]
        for s in streams[:4]:
            t = s.get("codec_type", "?")
            if t == "video":
                lines.append(f"Video stream: {s.get('codec_name')} · {s.get('width')}×{s.get('height')} · {s.get('r_frame_rate', '?')} fps")
            elif t == "audio":
                lines.append(f"Audio stream: {s.get('codec_name')} · {s.get('sample_rate')} Hz · {s.get('channels')} ch")
            else:
                lines.append(f"{t} stream: {s.get('codec_name')}")
        return "\n".join(lines)
    except Exception as e:
        return f"(ffprobe error: {e})"


def _preview_archive(path: Path, fmt: str) -> str:
    try:
        if fmt == "zip":
            with zipfile.ZipFile(str(path)) as z:
                names = z.namelist()
                return f"ZIP · {len(names)} entries\n" + "\n".join(names[:30])
        if fmt in ("tar", "tgz", "gz", "tbz2", "bz2", "txz", "xz"):
            import tarfile
            mode = "r" if fmt == "tar" else "r:*"
            try:
                with tarfile.open(str(path), mode) as t:
                    names = t.getnames()
                    return f"TAR · {len(names)} entries\n" + "\n".join(names[:30])
            except Exception:
                pass
        if fmt == "7z":
            import py7zr
            with py7zr.SevenZipFile(str(path), "r") as sz:
                names = sz.getnames()
                return f"7Z · {len(names)} entries\n" + "\n".join(names[:30])
        return f"(unsupported archive preview: {fmt})"
    except Exception as e:
        return f"(archive preview failed: {e})"


def _preview_ebook(path: Path, fmt: str) -> str:
    if fmt == "epub":
        try:
            import re as _re
            with zipfile.ZipFile(str(path)) as z:
                names = z.namelist()
                lines = [f"EPUB · {len(names)} entries"]
                # Grab title/author from the OPF manifest
                for n in names:
                    if n.endswith(".opf"):
                        try:
                            opf = z.read(n).decode("utf-8", "replace")
                            for tag in ("title", "creator", "language", "publisher", "subject"):
                                for m in _re.finditer(rf"<dc:{tag}[^>]*>([^<]+)</dc:{tag}>", opf):
                                    lines.append(f"{tag}: {m.group(1).strip()}")
                        except Exception:
                            pass
                        break
                # Grab a few paragraphs of actual reading text from the first HTML file
                html_names = [n for n in names if n.endswith((".xhtml", ".html", ".htm"))]
                for hn in html_names[:2]:
                    try:
                        raw = z.read(hn).decode("utf-8", "replace")
                        # Strip tags crudely
                        text = _re.sub(r"<[^>]+>", " ", raw)
                        text = _re.sub(r"\s+", " ", text).strip()
                        if text:
                            lines.append(f"--- {hn} ---")
                            lines.append(text[:1200])
                    except Exception:
                        pass
                return "\n".join(lines)[:MAX_PREVIEW_CHARS]
        except Exception as e:
            return f"(epub preview failed: {e})"
    return f"Ebook · {fmt.upper()} · {path.stat().st_size} bytes"


def _preview_font(path: Path) -> str:
    try:
        from fontTools.ttLib import TTFont
        tt = TTFont(str(path), lazy=True)
        name = tt.get("name")
        lines = ["Font metadata:"]
        if name:
            wanted = {1: "Family", 2: "Subfamily", 4: "Full name", 5: "Version", 6: "PostScript name"}
            for r in name.names:
                if r.nameID in wanted:
                    try:
                        lines.append(f"  {wanted[r.nameID]}: {r.toUnicode()}")
                    except Exception:
                        pass
        cmap = tt.getBestCmap()
        lines.append(f"Glyphs mapped: {len(cmap)}")
        lines.append(f"Tables: {', '.join(sorted(tt.keys()))}")
        tt.close()
        return "\n".join(lines)
    except Exception as e:
        return f"(font preview failed: {e})"


def extract_preview(path: Path, input_format: str, category: str) -> str:
    """Return a short human-readable preview of the file's contents."""
    fmt = input_format.lower()
    if category == "document":
        return _preview_document(path, fmt)
    if category == "image":
        return _preview_image(path)
    if category in ("audio", "video"):
        return _preview_media(path)
    if category == "archive":
        return _preview_archive(path, fmt)
    if category == "ebook":
        return _preview_ebook(path, fmt)
    if category == "font":
        return _preview_font(path)
    if category == "data":
        return _read_text(path)
    return _read_text(path)


# ---------- model call ----------

SYSTEM_PROMPT = """You are the AI assistant inside Ash Loves Files (ALF), a universal file converter.

Given a file's metadata and a content preview, produce a JSON response with these fields:

{
  "summary": "2-3 sentence plain-English description of what's in the file. Reference concrete details (topic, style, duration, dimensions, number of pages/items, genre, etc.) when visible.",
  "recommended": {
    "format": "one of the available_output_formats — the single best choice",
    "reason": "ONE short sentence (max 14 words) on why"
  },
  "alternatives": [
    {"format": "another available_output_format", "reason": "short sentence (max 14 words)"},
    {"format": "another one", "reason": "short sentence (max 14 words)"}
  ],
  "tips": [
    "one short gotcha or quality note the user should know (max 16 words)",
    "another short tip if relevant"
  ],
  "suggested_filename": "descriptive-lowercase-filename-without-extension"
}

Rules:
- You MUST pick formats strictly from the available_output_formats list.
- Recommendations should reflect the likely USE CASE (editing vs archiving vs sharing vs web vs print).
- Suggested filename should be kebab-case, descriptive, 2-6 words, NO extension.
- Respond with the JSON object only — no prose before or after, no markdown fences."""


def analyze(
    filename: str,
    input_format: str,
    file_size: int,
    category: str,
    available_output_formats: list[str],
    preview: str,
) -> dict:
    client = _get_client()
    if client is None:
        return {"error": "AI is disabled (no GROQ_API_KEY configured)"}

    user_msg = f"""FILENAME: {filename}
INPUT_FORMAT: {input_format}
CATEGORY: {category}
SIZE: {file_size} bytes
AVAILABLE_OUTPUT_FORMATS: {', '.join(available_output_formats)}

CONTENT PREVIEW:
{preview[:MAX_PREVIEW_CHARS]}
"""

    try:
        resp = client.chat.completions.create(
            model=settings.GROQ_MODEL,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_msg},
            ],
            temperature=0.4,
            max_tokens=700,
            response_format={"type": "json_object"},
        )
        raw = resp.choices[0].message.content or "{}"
        data = json.loads(raw)
    except Exception as e:
        log.warning("Groq call failed: %s", e)
        return {"error": f"AI analysis failed: {e}"}

    # Normalise / validate shape — clamp anything weird so the frontend can trust it
    allowed = {f.lower() for f in available_output_formats}

    def _safe_alt(item):
        if not isinstance(item, dict):
            return None
        fmt = str(item.get("format", "")).lower().lstrip(".")
        if fmt not in allowed:
            return None
        return {"format": fmt, "reason": str(item.get("reason", ""))[:200]}

    rec = data.get("recommended") or {}
    rec_fmt = str(rec.get("format", "")).lower().lstrip(".")
    if rec_fmt not in allowed:
        rec_fmt = next(iter(sorted(allowed)), "")

    alternatives = []
    for item in data.get("alternatives") or []:
        safe = _safe_alt(item)
        if safe and safe["format"] != rec_fmt and not any(a["format"] == safe["format"] for a in alternatives):
            alternatives.append(safe)
        if len(alternatives) >= 3:
            break

    tips_raw = data.get("tips") or []
    tips = [str(t)[:250] for t in tips_raw if isinstance(t, str) and t.strip()][:3]

    return {
        "summary": str(data.get("summary", ""))[:800],
        "recommended": {"format": rec_fmt, "reason": str(rec.get("reason", ""))[:200]},
        "alternatives": alternatives,
        "tips": tips,
        "suggested_filename": str(data.get("suggested_filename", ""))[:80].strip(),
        "model": settings.GROQ_MODEL,
    }
